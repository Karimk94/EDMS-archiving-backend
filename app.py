from flask import Flask, jsonify, request, Response, send_file, session, abort
from flask_cors import CORS
import db_connector
import wsdl_client
import logging
from waitress import serve
import os
import json
import math
from functools import wraps
from datetime import timedelta
import mimetypes
import csv
import io
import openpyxl

# --- Logging Setup ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
app = Flask(__name__)
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=60)
app.secret_key = os.getenv('FLASK_SECRET_KEY', 'a_very_secret_default_key_replace_me')  # Added default for safety
CORS(app, supports_credentials=True, resources={r"/api/*": {"origins": "*"}})

# --- Security Decorator ---
def editor_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session or session['user'].get('security_level') != 'Editor':
            logging.warning(f"Editor access denied for user: {session.get('user')}")
            abort(403)  # Forbidden
        return f(*args, **kwargs)

    return decorated_function

# --- Authentication Routes (for Archiving Frontend) ---
@app.route('/api/auth/pta-login', methods=['POST'])
def pta_login():
    """
    Handles login specifically for the PTA Archiving frontend.
    """
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')

    if not username or not password:
        return jsonify({"error": "Username and password are required"}), 400

    # Use the DMS user login function for authentication
    dst = wsdl_client.dms_user_login(username, password)

    if dst:
        # If DMS login is successful, get security level from our new table
        security_level = db_connector.get_pta_user_security_level(username)

        if security_level is None:
            # User exists in DMS but not in our security setup, or DB error
            logging.warning(
                f"User '{username}' authenticated via DMS but has no security level assigned in middleware DB.")
            return jsonify({"error": "User not authorized for this application"}), 401

        session['user'] = {'username': username, 'security_level': security_level}
        # We store the DST in the session for this user
        session['dst'] = dst
        session.permanent = True
        logging.info(f"User '{username}' logged in successfully with security level '{security_level}'.")
        return jsonify({"message": "Login successful", "user": session['user']}), 200
    else:
        logging.warning(f"DMS login failed for user '{username}'.")
        return jsonify({"error": "Invalid DMS credentials"}), 401

@app.route('/api/auth/pta-user', methods=['GET'])
def get_pta_user():
    """
    Gets the session for the PTA Archiving frontend user.
    """
    user_session = session.get('user')
    if user_session and 'username' in user_session:
        # Re-fetch details to ensure they are current
        user_details = db_connector.get_pta_user_details(user_session['username'])
        if user_details:
            session['user'] = user_details  # Update session
            return jsonify({'user': user_details}), 200
        else:
            # User was in session but is no longer valid
            session.pop('user', None)
            session.pop('dst', None)
            return jsonify({'error': 'User not found'}), 401
    else:
        return jsonify({'error': 'Not authenticated'}), 401

# --- Auth routes also used by `page.tsx` (generic, but frontend depends on them) ---
@app.route('/api/auth/logout', methods=['POST'])
def logout():
    username = session.get('user', {}).get('username', 'Unknown user')
    session.pop('user', None)
    session.pop('dst', None)  # Clear user's DMS session token
    logging.info(f"User '{username}' logged out.")
    return jsonify({"message": "Logout successful"}), 200

@app.route('/api/auth/pta-user', methods=['GET'])
def get_user():
    """
    Generic user check. The frontend `page.tsx` uses this.
    It relies on a different db_connector function, so we must include it.
    """
    user_session = session.get('user')
    if user_session and 'username' in user_session:
        user_details = db_connector.get_pta_user_details(user_session['username'])
        if user_details:
            session['user'] = user_details  # Update session
            return jsonify({'user': user_details}), 200
        else:
            # User was in session but not in DB? Log them out.
            session.pop('user', None)
            session.pop('dst', None)
            return jsonify({'error': 'User not found'}), 401
    else:
        return jsonify({'error': 'Not authenticated'}), 401

# --- Archiving API Routes ---
@app.route('/api/dashboard_counts', methods=['GET'])
def get_dashboard_counts():
    if 'user' not in session:
        return jsonify({"error": "Unauthorized"}), 401
    counts = db_connector.get_dashboard_counts()
    return jsonify(counts)

@app.route('/api/employees', methods=['GET'])
def get_employees():
    if 'user' not in session: return jsonify({"error": "Unauthorized"}), 401
    employees, total_rows = db_connector.fetch_archived_employees(
        page=request.args.get('page', 1, type=int),
        page_size=request.args.get('page_size', 20, type=int),
        search_term=request.args.get('search'),
        status=request.args.get('status'),
        filter_type=request.args.get('filter_type')
    )
    total_pages = math.ceil(total_rows / request.args.get('page_size', 20, type=int))
    return jsonify({"employees": employees, "total_employees": total_rows, "total_pages": total_pages})

@app.route('/api/employees', methods=['POST'])
@editor_required
def add_employee_archive():
    if 'user' not in session or 'dst' not in session:
        return jsonify({"error": "Unauthorized or session expired"}), 401

    try:
        # Use the user's DMS session token
        dst = session['dst']
        dms_user = session['user']['username']
        employee_data = json.loads(request.form.get('employee_data'))

        documents = []
        i = 0
        while f'new_documents[{i}][file]' in request.files:
            doc_data = {
                "file": request.files[f'new_documents[{i}][file]'],
                "doc_type_id": request.form.get(f'new_documents[{i}][doc_type_id]'),
                "doc_type_name": request.form.get(f'new_documents[{i}][doc_type_name]'),
                "expiry": request.form.get(f'new_documents[{i}][expiry]'),
                "legislation_ids": request.form.getlist(f'new_documents[{i}][legislation_ids][]')
            }
            documents.append(doc_data)
            i += 1

        # Pass the user's DST to the function
        success, message = db_connector.add_employee_archive_with_docs(dst, dms_user, employee_data, documents)

        return (jsonify({"message": message}), 201) if success else (jsonify({"error": message}), 500)

    except Exception as e:
        logging.error(f"Error adding employee archive: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route('/api/employees/<int:archive_id>', methods=['GET'])
def get_employee_details(archive_id):
    if 'user' not in session: return jsonify({"error": "Unauthorized"}), 401
    details = db_connector.fetch_single_archived_employee(archive_id)
    return jsonify(details) if details else (jsonify({"error": "Not found"}), 404)

@app.route('/api/employees/<int:archive_id>', methods=['PUT'])
@editor_required
def update_employee_archive(archive_id):
    if 'user' not in session or 'dst' not in session:
        return jsonify({"error": "Unauthorized or session expired"}), 401

    try:
        # Use the user's DMS session token
        dst = session['dst']
        dms_user = session['user']['username']
        employee_data = json.loads(request.form.get('employee_data'))

        new_documents = []
        i = 0
        while f'new_documents[{i}][file]' in request.files:
            doc_data = {
                "file": request.files[f'new_documents[{i}][file]'],
                "doc_type_id": request.form.get(f'new_documents[{i}][doc_type_id]'),
                "doc_type_name": request.form.get(f'new_documents[{i}][doc_type_name]'),
                "expiry": request.form.get(f'new_documents[{i}][expiry]'),
                "legislation_ids": request.form.getlist(f'new_documents[{i}][legislation_ids][]')
            }
            new_documents.append(doc_data)
            i += 1

        deleted_doc_ids = json.loads(request.form.get('deleted_documents', '[]'))
        updated_documents = json.loads(request.form.get('updated_documents', '[]'))

        success, message = db_connector.update_archived_employee(
            dst, dms_user, archive_id, employee_data,
            new_documents, deleted_doc_ids, updated_documents
        )

        return (jsonify({"message": message}), 200) if success else (jsonify({"error": message}), 500)
    except Exception as e:
        logging.error(f"Error updating employee archive {archive_id}: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route('/api/hr_employees', methods=['GET'])
def get_hr_employees():
    if 'user' not in session: return jsonify({"error": "Unauthorized"}), 401
    search = request.args.get('search', "", type=str)
    page = request.args.get('page', 1, type=int)
    employees, total_rows = db_connector.fetch_hr_employees_paginated(search_term=search, page=page)
    has_more = (page * 10) < total_rows
    return jsonify({"employees": employees, "hasMore": has_more})

@app.route('/api/hr_employees/<int:employee_id>', methods=['GET'])
def get_hr_employee_details(employee_id):
    if 'user' not in session: return jsonify({"error": "Unauthorized"}), 401
    details = db_connector.fetch_hr_employee_details(employee_id)
    return jsonify(details) if details else (jsonify({"error": "Not found"}), 404)

@app.route('/api/statuses', methods=['GET'])
def get_statuses():
    if 'user' not in session: return jsonify({"error": "Unauthorized"}), 401
    statuses = db_connector.fetch_statuses()
    return jsonify(statuses)

@app.route('/api/document_types', methods=['GET'])
def get_document_types():
    if 'user' not in session: return jsonify({"error": "Unauthorized"}), 401
    doc_types = db_connector.fetch_document_types()
    return jsonify(doc_types)

@app.route('/api/legislations', methods=['GET'])
def get_legislations():
    if 'user' not in session: return jsonify({"error": "Unauthorized"}), 401
    legislations = db_connector.fetch_legislations()
    return jsonify(legislations)

@app.route('/api/document/<int:docnumber>', methods=['GET'])
def get_document_file(docnumber):
    """
    Securely streams a document from the DMS to the client.
    It uses the user's session DST to authorize the download.
    """
    if 'user' not in session or 'dst' not in session:
        return jsonify({"error": "Unauthorized or session expired"}), 401

    # Use the user's DMS session token from their login
    dst = session['dst']

    file_bytes, filename = wsdl_client.get_document_from_dms(dst, docnumber)

    if file_bytes and filename:
        mimetype = mimetypes.guess_type(filename)[0] or 'application/octet-stream'
        return Response(file_bytes, mimetype=mimetype, headers={"Content-Disposition": f"inline; filename={filename}"})
    else:
        logging.warning(f"Document not found or retrieval failed for docnumber: {docnumber}")
        return jsonify({"error": "Document not found or could not be retrieved from DMS."}), 404


@app.route('/api/employees/bulk-upload', methods=['POST'])
@editor_required
def bulk_upload_employees():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    employees_data = []

    try:
        # Define expected columns based on the image/CSV from previous step
        expected_headers = [
            "Employee ID", "Name (AR)", "Name (EN)", "Hire Date",
            "Nationality", "Job Title", "Manager", "Phone", "Email",
            "Employee Status", "Section", "Department"
        ]

        if file.filename.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file.stream)
            sheet = workbook.active

            # Read header
            header_cells = next(sheet.iter_rows(min_row=1, max_row=1))
            header = [cell.value for cell in header_cells]

            # Basic header validation
            if not all(h in header for h in expected_headers):
                return jsonify({
                                   "error": f"Invalid Excel format. Missing one or more headers. Expected: {', '.join(expected_headers)}"}), 400

            header_map = {h: i for i, h in enumerate(header)}

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(c is None for c in row):  # Skip empty rows
                    continue

                # Handle different date formats from Excel
                hire_date_val = row[header_map["Hire Date"]]
                hire_date_str = None
                if isinstance(hire_date_val, datetime):
                    hire_date_str = hire_date_val.strftime('%d/%m/%Y')
                elif isinstance(hire_date_val, str):
                    hire_date_str = hire_date_val  # Assume it's in DD/MM/YYYY format

                emp_data = {
                    "empno": row[header_map["Employee ID"]],
                    "name_ar": row[header_map["Name (AR)"]],
                    "name_en": row[header_map["Name (EN)"]],
                    "hire_date": hire_date_str,
                    "nationality": row[header_map["Nationality"]],
                    "job_title": row[header_map["Job Title"]],
                    "manager": row[header_map["Manager"]],
                    "phone": row[header_map["Phone"]],
                    "email": row[header_map["Email"]],
                    "status_name": row[header_map["Employee Status"]],
                    "section": row[header_map["Section"]],
                    "department": row[header_map["Department"]]
                }
                employees_data.append(emp_data)

        elif file.filename.endswith('.csv'):
            stream = io.StringIO(file.stream.read().decode("utf-8-sig"), newline=None)  # Use utf-8-sig to handle BOM
            csv_reader = csv.reader(stream)

            header = next(csv_reader)

            if not all(h in header for h in expected_headers):
                return jsonify({
                                   "error": f"Invalid CSV format. Missing one or more headers. Expected: {', '.join(expected_headers)}"}), 400

            header_map = {h: i for i, h in enumerate(header)}

            for row in csv_reader:
                if not any(row):  # Skip empty rows
                    continue

                emp_data = {
                    "empno": row[header_map["Employee ID"]],
                    "name_ar": row[header_map["Name (AR)"]],
                    "name_en": row[header_map["Name (EN)"]],
                    "hire_date": row[header_map["Hire Date"]],  # Assume DD/MM/YYYY string
                    "nationality": row[header_map["Nationality"]],
                    "job_title": row[header_map["Job Title"]],
                    "manager": row[header_map["Manager"]],
                    "phone": row[header_map["Phone"]],
                    "email": row[header_map["Email"]],
                    "status_name": row[header_map["Employee Status"]],
                    "section": row[header_map["Section"]],
                    "department": row[header_map["Department"]]
                }
                employees_data.append(emp_data)

        else:
            return jsonify({"error": "Invalid file type. Please upload a .xlsx or .csv file."}), 400

        if not employees_data:
            return jsonify({"error": "No data found in the file."}), 400

        # Process the data
        success, failed, errors = db_connector.bulk_add_employees_from_excel(employees_data)

        if failed > 0:
            return jsonify({
                "message": f"Bulk add finished. {success} added, {failed} failed.",
                "errors": errors
            }), 422  # Unprocessable Entity
        else:
            return jsonify({
                "message": f"Successfully added {success} employees."
            }), 201

    except Exception as e:
        logging.error(f"Error processing bulk upload file: {e}", exc_info=True)
        return jsonify({"error": f"An error occurred during file processing: {str(e)}"}), 500

@app.route('/api/employees/export', methods=['GET'])
def export_employees():
    if 'user' not in session:
        return jsonify({"error": "Unauthorized"}), 401

    try:
        # 1. Get filter parameters from the request, same as the dashboard
        search_term = request.args.get('search')
        status = request.args.get('status')
        filter_type = request.args.get('filter_type')

        # 2. Fetch *all* matching employees. We set a very high page_size.
        employees, total_rows = db_connector.fetch_archived_employees(
            page=1,
            page_size=0,
            search_term=search_term,
            status=status,
            filter_type=filter_type
        )

        if not employees:
            return jsonify({"error": "No data to export for this filter"}), 404

        # 3. Define CSV headers based on the dashboard table
        headers = [
            "EmpNo", "FullName_EN", "FullName_AR", "Department", "Section",
            "Status_EN", "Status_AR", "Warrant_Status", "Card_Status", "Card_Expiry"
        ]

        # 4. Create CSV in-memory
        si = io.StringIO()
        cw = csv.writer(si)

        # Write header
        cw.writerow(headers)

        # Write employee data rows
        for emp in employees:
            cw.writerow([
                emp.get('empno'),
                emp.get('fullname_en'),
                emp.get('fullname_ar'),
                emp.get('department'),
                emp.get('section'),
                emp.get('status_en'),
                emp.get('status_ar'),
                emp.get('warrant_status'),
                emp.get('card_status'),
                emp.get('card_expiry')
            ])

        # 5. Prepare and return the CSV file as a response
        output = si.getvalue()
        si.close()

        return Response(
            output,
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment;filename=employee_export.csv"}
        )

    except Exception as e:
        logging.error(f"Error exporting employees: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('HTTP_PLATFORM_PORT', 5006))
    logging.info(f"Starting Migrated Archiving Backend on host 0.0.0.0 port {port}")
    serve(app, host='0.0.0.0', port=port, threads=50)